# -*- coding: UTF-8 -*-

import sys
import time
import urllib
import urllib2
# import requests
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook

reload(sys)
sys.setdefaultencoding('utf8')

#Some User Agents
hds=[{'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},\
{'User-Agent':'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},\
{'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]


########################################################################################################################
def book_spider(book_tag):
    page_num = 0
    book_list = []
    try_times = 0

    while (1):
        # url='http://www.douban.com/tag/%E5%B0%8F%E8%AF%B4/book?start=0' # For Test

        # 一页显示15本书 所以乘以15
        # https://book.douban.com/people/49754936/wish?start=0&sort=time&rating=all&filter=all&mode=grid
        # https://book.douban.com/people/49754936/wish?start=15&sort=time&rating=all&filter=all&mode=grid
        # https://book.douban.com/people/49754936/wish?start=0 同样可以访问 减少 url 长度

        url = 'https://www.douban.com/tag/' + urllib.quote(book_tag) + '/book?start=' + str(page_num * 15)
        time.sleep(np.random.rand() * 5)

        # Last Version
        try:
            req = urllib2.Request(url, headers=hds[page_num % len(hds)])
            source_code = urllib2.urlopen(req).read()
            plain_text = str(source_code)
        except (urllib2.HTTPError, urllib2.URLError), e:
            print e
            continue

        # Previous Version, IP is easy to be Forbidden
        # source_code = requests.get(url)
        # plain_text = source_code.text

        # 通过 BeautifulSoup 获取 HTML中的数据
        soup = BeautifulSoup(plain_text)
        # 查询
        list_soup = soup.find('div', {'class': 'mod book-list'})

        # 尝试200次读取信息
        try_times += 1
        if list_soup == None and try_times < 200:
            continue
        elif list_soup == None or len(list_soup) <= 1:
            break  # Break when no informatoin got after 200 times requesting

        # 查找HTML source code 中的 所有 dd 标签的内容
        for book_info in list_soup.findAll('dd'):
            # 获取书名
            title = book_info.find('a', {'class': 'title'}).string.strip()
            # 获取出版社作者
            desc = book_info.find('div', {'class': 'desc'}).string.strip()
            desc_list = desc.split('/')
            book_url = book_info.find('a', {'class': 'title'}).get('href')
            # 例： [美]安德斯·艾利克森（AndersEricsson）、罗伯特·普尔（RobertPool） / 王正林 / 机械工业出版社 / 2016 - 11 - 6 / 39.00元
            try:
                author_info = '作者/译者： ' + '/'.join(desc_list[0:-3])
            except:
                author_info = '作者/译者： 暂无'
            try:
                pub_info = '出版信息： ' + '/'.join(desc_list[-3:])
            except:
                pub_info = '出版信息： 暂无'
            try:
                rating = book_info.find('span', {'class': 'rating_nums'}).string.strip()
            except:
                rating = '0.0'
            try:
                # people_num = book_info.findAll('span')[2].string.strip()
                people_num = get_people_num(book_url)
                people_num = people_num.strip('人评价')
            except:
                people_num = '0'

            # 添加书本信息到book list
            book_list.append([title, rating, people_num, author_info, pub_info])
            try_times = 0  # set 0 when got valid information
        page_num += 1
        print 'Downloading Information From Page %d' % page_num
    return book_list


def get_people_num(url):
    # url='http://book.douban.com/subject/6082808/?from=tag_all' # For Test
    try:
        # 构造request
        req = urllib2.Request(url, headers=hds[np.random.randint(0, len(hds))])
        # 获取 replay 中的 source code
        source_code = urllib2.urlopen(req).read()
        # 解析 source code
        plain_text = str(source_code)
    except (urllib2.HTTPError, urllib2.URLError), e:
        # 报错处理
        plain_text = []
        print e
    # 通过 BeautifulSoup 获取 HTML中的数据
    soup = BeautifulSoup(plain_text)
    # 查询
    people_num = soup.find('div', {'class': 'rating_sum'}).findAll('span')[1].string.strip()
    return people_num


def do_spider(book_tag_lists):
    # 创建book_lists空数组
    book_lists = []
    # 遍历book_tag
    for book_tag in book_tag_lists:
        # 每一个book tag执行爬虫
        book_list = book_spider(book_tag)
        # 爬虫结果排序
        book_list = sorted(book_list, key=lambda x: x[1], reverse=True)
        # 将排序结果保存在book_lists中， 注意和book_list的区别
        book_lists.append(book_list)
    return book_lists


def print_book_lists_excel(book_lists, book_tag_lists):
    wb = Workbook(optimized_write=True)
    ws = []
    # 更具tag创建各自的子表，就是excel最下面的sheets
    for i in range(len(book_tag_lists)):
        ws.append(wb.create_sheet(title=book_tag_lists[i].decode()))  # utf8->unicode

    # 遍历每一个tag的对应的sheet
    for i in range(len(book_tag_lists)):

        # 在tag对应的sheet中添加第一行，表明每列的内容
        ws[i].append(['序号', '书名', '评分', '评价人数', '作者', '出版社'])
        count = 1

        # 遍历book_lists这个数组，序号和book_tag_lists对应
        for bl in book_lists[i]:
            #             序号   书名   评分         评价人数     作者   出版社
            ws[i].append([count, bl[0], float(bl[1]), int(bl[2]), bl[3], bl[4]])
            count += 1

    # 设置保存文件的文件头字符串 参考：book_list-个人管理-时间管理-投资-文化-宗教.xlsx
    save_path = 'book_list'

    # 设置保存文件的文件中tag的描述 参考：book_list-个人管理-时间管理-投资-文化-宗教.xlsx
    for i in range(len(book_tag_lists)):
        save_path += ('-' + book_tag_lists[i].decode())

    # 设置保存文件的文件类型 参考：book_list-个人管理-时间管理-投资-文化-宗教.xlsx
    save_path += '.xlsx'

    # 保存文件
    wb.save(save_path)


########################################################################################################################
def book_status_to_url(book_status):
    url = {
        '正在看': 'do',
        '看过': 'collect',
        '想看': 'wish',
    }
    return url.get(book_status, None)


# def book_spider_with_user_id_and_status(user_id, book_status):
def book_spider_with_user_id_and_status(book_status):
    page_num = 0
    book_list = []
    try_times = 0

    while 1:
        # 一页显示15本书 所以乘以15
        # https://book.douban.com/people/49754936/do?start=0
        # https://book.douban.com/people/49754936/wish?start=0
        # https://book.douban.com/people/49754936/collect?start=0

        # url = 'https://book.douban.com/people/49754936/collect?start=0'  # For Test
        url = 'https://book.douban.com/people/' + user_id + '/' + book_status_to_url(book_status) + '?start=' + str(page_num * 15)

        headers = {'User-Agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.93 Safari/537.36',
                   'Cookie': 'bid=ml0q26zGHcs; ll="108169"; ap_v=0,6.0; __utma=30149280.58262688.1620584802.1620584802.1620584802.1; __utmc=30149280; __utmz=30149280.1620584802.1.1.utmcsr=(direct)|utmccn=(direct)|utmcmd=(none); __utmt_douban=1; __utma=81379588.1453940470.1620584802.1620584802.1620584802.1; __utmc=81379588; __utmz=81379588.1620584802.1.1.utmcsr=(direct)|utmccn=(direct)|utmcmd=(none); __utmt=1; _pk_ses.100001.3ac3=*; _pk_id.100001.3ac3=15033d4ce8d72d6b.1620584802.1.1620585128.1620584802.; __utmb=30149280.6.10.1620584802; __utmb=81379588.6.10.1620584802'}

        print "URL is:", url
        print "Random is:", np.random.random()

        time.sleep(np.random.random())

        # Last Version
        try:
            # req = urllib2.Request(url, headers=hds[page_num % len(hds)])
            req = urllib2.Request(url, headers=headers)
            source_code = urllib2.urlopen(req).read()
            plain_text = str(source_code)
        except (urllib2.HTTPError, urllib2.URLError), e:
            print e
            continue

        # Previous Version, IP is easy to be Forbidden
        # source_code = requests.get(url)
        # plain_text = source_code.text

        # 通过 BeautifulSoup 获取 HTML中的数据
        soup = BeautifulSoup(plain_text,features="html.parser")
        # 查询
        list_soup = soup.find('ul', {'class': 'interest-list'})

        # 尝试200次读取信息
        try_times += 1
        if list_soup is None and try_times < 200:
            continue
        elif list_soup is None or len(list_soup) <= 1:
            break  # Break when no information got after 200 times requesting

        book_info_lists = list_soup.findAll('li', {'class': 'subject-item'})

        for book_info in book_info_lists:
        # for book_info in list_soup.findAll('li', {'class': 'subject-item'}):
            sub_book_info = book_info.find('div', {'class': 'info'})
            title = sub_book_info.find('a').string.strip()
            # print "Title is:", title
            # 获取出版社作者
            desc = sub_book_info.find('div', {'class': 'pub'}).string.strip()
            # print "Desc is:", desc
            desc_list = desc.split('/')
            # 例： [美]安德斯·艾利克森（AndersEricsson）、罗伯特·普尔（RobertPool） / 王正林 / 机械工业出版社 / 2016 - 11 - 6 / 39.00元
            try:
                author_info = '作者/译者： ' + '/'.join(desc_list[0:-3])
                # print "Author is:", author_info
            except:
                author_info = '作者/译者： 暂无'
            try:
                pub_info = '出版信息： ' + '/'.join(desc_list[-3:])
                # print "Public is:", pub_info
            except:
                pub_info = '出版信息： 暂无'

            print "Author is:", author_info
            print "Public is:", pub_info

            # 添加书本信息到book list
            book_list.append([title, author_info, pub_info])
            try_times = 0  # set 0 when got valid information
        page_num += 1
        print 'Downloading Information From Page %d' % page_num
    return book_list


def do_spider_with_user_id(user_id):
    # 创建book_lists空数组
    book_lists = []
    # book_status_lists = ['正在看', '看过', '想看']
    book_status_lists = ['看过']
    # 遍历book_tag
    for book_status in book_status_lists:
        # 每一个book status执行爬虫
        # book_list = book_spider_with_user_id_and_status(user_id, book_status)
        book_list = book_spider_with_user_id_and_status(book_status)
        # 爬虫结果排序
        book_list = sorted(book_list, key=lambda x: x[1], reverse=True)
        # 将排序结果保存在book_lists中， 注意和book_list的区别
        book_lists.append(book_list)
    return book_lists


def print_book_lists_excel_with_user_id(book_lists):
    # book_status_lists = ['正在看', '看过', '想看']
    book_status_lists = ['看过']

    wb = Workbook(optimized_write=True)
    ws = []
    # 更具tag创建各自的子表，就是excel最下面的sheets
    for i in range(len(book_status_lists)):
        ws.append(wb.create_sheet(title=book_lists[i].decode()))  # utf8->unicode

    # 遍历每一个status的对应的sheet
    for i in range(len(book_status_lists)):

        # 在tag对应的sheet中添加第一行，表明每列的内容
        ws[i].append(['序号', '书名', '作者', '出版社'])
        count = 1

        # 遍历book_lists这个数组，序号和book_tag_lists对应
        for bl in book_lists[i]:
            #             序号    书名   作者    出版社
            ws[i].append([count, bl[0], bl[3], bl[4]])
            count += 1

    # 设置保存文件的文件头字符串 参考：book_list-个人管理-时间管理-投资-文化-宗教.xlsx
    save_path = 'book_list'

    # 设置保存文件的文件中tag的描述 参考：book_list-个人管理-时间管理-投资-文化-宗教.xlsx
    for i in range(len(book_status_lists)):
        save_path += ('-' + book_status_lists[i].decode())

    # 设置保存文件的文件类型 参考：book_list-个人管理-时间管理-投资-文化-宗教.xlsx
    save_path += '.xlsx'

    # 保存文件
    wb.save(save_path)


########################################################################################################################

if __name__ == '__main__':
    # 按照标签爬取数据

    # book_tag_lists = ['心理','判断与决策','算法','数据结构','经济','历史']
    # book_tag_lists = ['传记','哲学','编程','创业','理财','社会学','佛教']
    # book_tag_lists = ['思想','科技','科学','web','股票','爱情','两性']
    # book_tag_lists = ['计算机','机器学习','linux','android','数据库','互联网']
    # book_tag_lists = ['数学']
    # book_tag_lists = ['摄影','设计','音乐','旅行','教育','成长','情感','育儿','健康','养生']
    # book_tag_lists = ['商业','理财','管理']
    # book_tag_lists = ['名著']
    # book_tag_lists = ['科普','经典','生活','心灵','文学']
    # book_tag_lists = ['科幻','思维','金融']
    # book_tag_lists = ['个人管理', '时间管理', '投资', '文化', '宗教']
    # book_lists = do_spider(book_tag_lists)
    # print_book_lists_excel(book_lists, book_tag_lists)

    # 按照用户爬取数据
    user_id = '49754936'
    book_lists = do_spider_with_user_id(user_id)
    print_book_lists_excel_with_user_id(book_lists)
